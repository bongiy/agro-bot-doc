import sys
import pathlib
import openpyxl

sys.path.append(str(pathlib.Path(__file__).resolve().parents[1]))
from utils.company_report import company_report_to_excel


def test_company_report_to_excel_generates_workbook():
    summary = [
        {
            "name": "ТОВ 'Зоря'",
            "contracts": 2,
            "plots": 3,
            "physical_area": 10.0,
            "contract_area": 9.5,
            "payers": 4,
            "rent_total": 1000,
            "paid_total": 800,
        }
    ]
    types = [
        {"company_name": "ТОВ 'Зоря'", "contract_type": "lease", "count": 2, "area": 9.5}
    ]
    sublease = [
        {
            "company_name": "ТОВ 'Зоря'",
            "received_plots": 1,
            "received_area": 2.0,
            "transferred_plots": 0,
            "transferred_area": 0.0,
        }
    ]
    payments = [
        {"company": "ТОВ 'Зоря'", "year": 2024, "accrued": 1000, "paid": 800, "debt": 200}
    ]
    bio = company_report_to_excel(summary, types, sublease, payments)
    wb = openpyxl.load_workbook(bio)
    assert "ТОВ" in wb.sheetnames
    assert "Типи договорів" in wb.sheetnames
    assert "Суборенда" in wb.sheetnames
    assert "Виплати по роках" in wb.sheetnames
    sheet = wb["ТОВ"]
    assert sheet.max_row == 2
    assert sheet["A2"].value == "ТОВ 'Зоря'"
