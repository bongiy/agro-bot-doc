from io import BytesIO
from typing import Iterable


def company_report_to_excel(summary: Iterable[dict], types: Iterable[dict], sublease: Iterable[dict], payments: Iterable[dict]) -> BytesIO:
    """Generate Excel workbook for company analytics report."""
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: summary
    ws = wb.active
    ws.title = "ТОВ"
    ws.append([
        "ТОВ",
        "Договорів",
        "Ділянок",
        "Фізична площа",
        "Площа за договорами",
        "Пайовиків",
        "Оренда",
        "Виплачено",
        "Покрито %",
    ])
    for r in summary:
        rent = float(r.get("rent_total") or 0)
        paid = float(r.get("paid_total") or 0)
        coverage = paid / rent * 100 if rent else 0
        ws.append([
            r.get("name"),
            r.get("contracts"),
            r.get("plots"),
            float(r.get("physical_area") or 0),
            float(r.get("contract_area") or 0),
            r.get("payers"),
            rent,
            paid,
            coverage,
        ])
        row = ws.max_row
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "0.00"
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=8).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "0.00"

    # Sheet 2: contract types
    type_labels = {
        "lease": ("Оренда (шт)", "Оренда (га)"),
        "short": ("Короткостр. (шт)", "Короткостр. (га)"),
        "emphyteusis": ("Емфітевзис (шт)", "Емфітевзис (га)"),
        "superficies": ("Суперфіцій (шт)", "Суперфіцій (га)"),
    }
    ws = wb.create_sheet("Типи договорів")
    header = ["ТОВ"]
    for t in type_labels.values():
        header.extend(t)
    ws.append(header)
    # Build pivot dictionary
    pivot: dict[str, dict[str, dict[str, float]]] = {}
    for r in types:
        comp = pivot.setdefault(r["company_name"], {k: {"count": 0, "area": 0.0} for k in type_labels})
        t = r["contract_type"]
        if t in comp:
            comp[t]["count"] = r["count"]
            comp[t]["area"] = float(r["area"] or 0)
    for comp_name in sorted(pivot.keys()):
        row = [comp_name]
        data = pivot[comp_name]
        for key in type_labels:
            row.append(data[key]["count"])
            row.append(data[key]["area"])
        ws.append(row)
        r_idx = ws.max_row
        ws.cell(row=r_idx, column=3).number_format = "0.00"
        ws.cell(row=r_idx, column=5).number_format = "0.00"
        ws.cell(row=r_idx, column=7).number_format = "0.00"
        ws.cell(row=r_idx, column=9).number_format = "0.00"

    # Sheet 3: sublease
    ws = wb.create_sheet("Суборенда")
    ws.append(["ТОВ", "Отримано (шт)", "Отримано (га)", "Передано (шт)", "Передано (га)"])
    for r in sublease:
        ws.append([
            r["company_name"],
            r.get("received_plots") or 0,
            float(r.get("received_area") or 0),
            r.get("transferred_plots") or 0,
            float(r.get("transferred_area") or 0),
        ])
        row = ws.max_row
        ws.cell(row=row, column=3).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "0.00"

    # Sheet 4: payments by year
    ws = wb.create_sheet("Виплати по роках")
    ws.append(["ТОВ", "Рік", "Начислено", "Виплачено", "Борг"])
    for r in payments:
        ws.append([
            r["company"],
            r["year"],
            float(r.get("accrued") or 0),
            float(r.get("paid") or 0),
            float(r.get("debt") or 0),
        ])
        row = ws.max_row
        ws.cell(row=row, column=3).number_format = "0.00"
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "0.00"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
