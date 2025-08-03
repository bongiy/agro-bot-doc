from io import BytesIO
from typing import Iterable


async def payments_to_excel(rows: Iterable[dict]) -> BytesIO:
    """Generate Excel file from payment rows and return as BytesIO."""
    # Import here to avoid hard dependency during module import
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Дата", "Пайовик", "Компанія", "Сума", "Статус", "Спадкоємець"])
    total = 0.0
    for r in rows:
        amount = float(r.get("amount") or 0)
        total += amount
        date_val = (
            r.get("payment_date").strftime("%d.%m.%Y")
            if r.get("payment_date")
            else ""
        )
        ws.append(
            [
                date_val,
                r.get("payer_name") or "",
                r.get("company_name") or "",
                amount,
                r.get("status") or "",
                "так" if r.get("is_heir") else "ні",
            ]
        )
        # Set currency format for the last added amount cell
        ws.cell(row=ws.max_row, column=4).number_format = "0.00"
    # Summary row
    ws.append(["", "", "Разом", total, "", ""])
    ws.cell(row=ws.max_row, column=4).number_format = "0.00"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


async def land_report_to_excel(rows: Iterable[dict]) -> BytesIO:
    """Generate Excel file for land plots report."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Кадастровий",
            "Площа",
            "НГО",
            "Пайовик",
            "Договір",
            "Орендар",
            "Дата закінчення",
            "Поле",
            "Річна орендна плата",
        ]
    )
    total_area = 0.0
    total_rent = 0.0
    for r in rows:
        area = float(r.get("area") or 0)
        ngo = float(r.get("ngo") or 0)
        rent = float(r.get("rent_amount") or 0)
        total_area += area
        total_rent += rent
        date_val = (
            r.get("date_valid_to").strftime("%d.%m.%Y")
            if r.get("date_valid_to")
            else ""
        )
        ws.append(
            [
                r.get("cadaster") or "",
                area,
                ngo,
                r.get("payer_name") or "",
                r.get("contract_number") or "",
                r.get("company_name") or "",
                date_val,
                r.get("field_name") or "",
                rent,
            ]
        )
        row = ws.max_row
        ws.cell(row=row, column=2).number_format = "0.0000"
        ws.cell(row=row, column=3).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "0.00"
    ws.append(["Разом", total_area, "", "", "", "", "", "", total_rent])
    last_row = ws.max_row
    ws.cell(row=last_row, column=2).number_format = "0.0000"
    ws.cell(row=last_row, column=9).number_format = "0.00"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


async def rent_summary_to_excel(rows: Iterable[dict]) -> BytesIO:
    """Generate Excel file for rent summary report."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Компанія",
            "Договорів",
            "Пайовиків",
            "Ділянок",
            "Нараховано",
            "Виплачено",
            "Борг",
            "Очікує",
            "Частково",
            "Оплачено",
        ]
    )
    totals = [0] * 6  # rent_total, paid_total, debt, pending, partial, paid
    for r in rows:
        debt = float(r.get("rent_total", 0)) - float(r.get("paid_total", 0))
        ws.append(
            [
                r.get("name"),
                r.get("contracts"),
                r.get("payers"),
                r.get("plots"),
                float(r.get("rent_total", 0)),
                float(r.get("paid_total", 0)),
                debt,
                float(r.get("pending_amount", 0)),
                float(r.get("partial_amount", 0)),
                float(r.get("paid_amount", 0)),
            ]
        )
        row = ws.max_row
        for col in range(5, 11):
            ws.cell(row=row, column=col).number_format = "0.00"
        totals[0] += float(r.get("rent_total", 0))
        totals[1] += float(r.get("paid_total", 0))
        totals[2] += debt
        totals[3] += float(r.get("pending_amount", 0))
        totals[4] += float(r.get("partial_amount", 0))
        totals[5] += float(r.get("paid_amount", 0))
    ws.append(
        [
            "Разом",
            "",
            "",
            "",
            *totals,
        ]
    )
    last_row = ws.max_row
    for col in range(5, 11):
        ws.cell(row=last_row, column=col).number_format = "0.00"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


async def land_overview_to_excel(
    summary: dict, fields: Iterable[dict], companies: Iterable[dict], statuses: Iterable[dict]
) -> BytesIO:
    """Generate Excel file for land overview report."""
    from openpyxl import Workbook

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Сумарно"
    ws.append(["Показник", "Значення"])
    ws.append(["Кількість ділянок", summary["plots"]])
    ws.append(["Загальна площа (га)", summary["area"]])
    ws.append(["Загальна НГО", summary["ngo"]])
    ws.append(["Унікальних пайовиків", summary["payers"]])
    ws.append(["Активних договорів", summary["contracts"]])
    ws.append(["ТОВ-орендарів", summary["companies"]])
    ws.cell(row=2, column=2).number_format = "0"
    ws.cell(row=3, column=2).number_format = "0.00"
    ws.cell(row=4, column=2).number_format = "0.00"

    # Fields sheet
    ws = wb.create_sheet("По полях")
    ws.append(["Поле", "Ділянок", "Площа (га)"])
    for f in fields:
        ws.append([f["name"], f["plots"], float(f["area"] or 0)])
        ws.cell(row=ws.max_row, column=3).number_format = "0.00"

    # Companies sheet
    ws = wb.create_sheet("По компаніях")
    ws.append(["Компанія", "Ділянок", "Площа (га)"])
    for c in companies:
        ws.append([c["name"], c["plots"], float(c["area"] or 0)])
        ws.cell(row=ws.max_row, column=3).number_format = "0.00"

    # Statuses sheet
    ws = wb.create_sheet("По статусах")
    ws.append(["Статус", "Ділянок", "Площа (га)"])
    for s in statuses:
        label = "З договором" if s["status"] == "with_contract" else "Без договору"
        ws.append([label, s["plots"], float(s["area"] or 0)])
        ws.cell(row=ws.max_row, column=3).number_format = "0.00"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
